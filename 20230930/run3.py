import openpyxl

def get_data(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.worksheets[0]
    price_list = {}
    
    for item in sheet.iter_rows(values_only=True):
        price_list[item[0]] = item[1]

    return price_list

#price_list = get_data(r"C:\Users\13059\Desktop\测试\价格表.xlsx")
#print(price_list)

def get_data_2(file_name, start_row, end_row, start_col, end_col):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.worksheets[0]
    price_list = {}

    for item in sheet.iter_rows(min_row= start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
        price_list[item[0]] = item[1]

    total = sheet.cell(end_row + 1, end_col -1).value

    return price_list, total

#date_list = get_data_2(r"C:\Users\13059\Desktop\测试\销售订单9-28货拉拉06 的副本.xlsx", 2, 12, 12, 13)
#print(date_list)

